import numpy as np
from pdos_helper import dos_grid_general,get_alpha,get_ao_ind, get_ind_ao_underc
from matplotlib import pyplot as plt
from matplotlib.ticker import MultipleLocator
import random
import pandas as pd
import sys
import math
import scipy.special
from openpyxl import Workbook, load_workbook
from scipy import stats

#this script is meant to extract the average value of a single QD parameter from the defect catalog for dots of a certain type
#this is probably more of a hard-coding sort of situation but I can pretend

file=sys.argv[1] #ideally the defect catalog itself, in my dropbox
measure=sys.argv[2] #do you want to average a metric ("av"), sum ("sum"), or plot a histogram ("hist")
column=sys.argv[3] #the column index you want to target
constrained=False
if len(sys.argv)>4:
	constraint_column=sys.argv[4] #the column of the constraint you want to target. if blank, all dots for which column is defined
	constraint=sys.argv[5] #the value the constraint should have. if blank, all dots for which column is defined
	constrained=True


#read in the defect catalog
wb = load_workbook(filename=file,data_only=True)
ws = wb.active


#so the trick for this is going to be extracting the value for each dot (for which the value is defined)
dot_row=0
dataset=[]
for i in range(2000):
	if ws["A"+str(i+1)].value: #how we seperate different dots

		if ws["A"+str(i+1)].value!="Small":
			#if ws["I"+str(i+1)].value=="Defect": #some manual additional constraints, because that's easier
				#if ws["M"+str(i+1)].value>0:

					dot_row=dot_row+1
					if dot_row==1: #once per dot

						if ws[column+str(i+1)].value and (not isinstance(ws[column+str(i+1)].value,str)): #if parameter is defined and is a number

							if constrained:
								if ws[constraint_column+str(i+1)].value==constraint: #if constraint
										dataset.append(ws[column+str(i+1)].value)
							else:
								dataset.append(ws[column+str(i+1)].value)

	else:
		dot_row=0

dataset=np.array(dataset)
if measure=="av" or measure=="average":
	print("The average value of",ws[column+str(1)].value, "is",np.average(dataset))
elif measure=="sum":
	print("The sum of",ws[column+str(1)].value, "is",np.sum(dataset))
elif measure=="hist":
	plt.hist(dataset,bins=[-.4,-.35,-0.3,-0.25,-0.2,-.15,-.1,-.05,0,0.05,.1,.15,.2,.25,.3,.35,.4])
	plt.title(ws[column+"1"].value)
	minor_locator = MultipleLocator(base=0.05)
	plt.gca().xaxis.set_minor_locator(minor_locator)
	#plt.hist(dataset)
	plt.show()
else:
	print("Second argument should be the measurement type: 'av' for average, 'sum' for sum, 'hist' for a histogram")
