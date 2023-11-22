import numpy as np
from pdos_helper import dos_grid_general,get_alpha,get_ao_ind, get_ind_ao_underc
from matplotlib import pyplot as plt
from matplotlib.ticker import MultipleLocator
import random
import pandas as pd
import sys
import math
import scipy.special
from openpyxl import Workbook, load_workbook
from scipy import stats

#a general code for making nice matplot plots of data from excel spreadsheets

#for now i'm just going to implement line charts but it would be nice to do others eventually

#it's so much easier to just do this for continuous chunks of data. don't @ me
#also, because i am lazy, this will only work for columns up to AZ and won't work if the column range includes both Z and AA

excel=sys.argv[1] #the name of the excel file
sheet=sys.argv[2] #the name of the sheet you are interested in
plot_title=sys.argv[3] #the title of the plot, in ""
plot_start=sys.argv[4] #the upper-right-most index of the data you want to plot. Includes column labels
plot_end=sys.argv[5] #the bottom-left-most index of the data you want to plot

#read in the data
wb = load_workbook(filename=excel,data_only=True)
ws=wb[sheet]

#extract target data into the corresponding array
alphabet="ABCDEFGHIJKLMNOPQRSTUVWXYZ" #lol

#parse start and end indeces
col_start=""
row_start=""
for i,char in enumerate(plot_start):
	if char.isalpha():
		col_start=col_start+char
	else:
		row_start=row_start+char
row_start=int(row_start)

col_end=""
row_end=""
for i,char in enumerate(plot_end):
	if char.isalpha():
		col_end=col_end+char
	else:
		row_end=row_end+char
row_end=int(row_end)

#how many lines will we need in our chart?
if len(col_end)==1 and len(col_end)==1:
	col_diff=1+alphabet.index(col_end)-alphabet.index(col_start)
elif len(col_end)==2 and len(col_end)==2:
	if col_start[0]=="B" or col_end[0]=="B":
		print("hey chill out i didn't think you'd get to the BX columns")
	new_alphabet=""
	for i,char in enumerate(alphabet):
		new_alphabet=new_alphabet+"A"
		new_alphabet=new_alphabet+char
	col_diff=1+new_alphabet.index(col_end)-new_alphabet.index(col_start)
else:
	print("You have to avoid the Z-AA column crossover because i'm too lazy to code around it")

#actually loop through
data=[]
series_names=[]
row_diff=1+row_end-row_start
for i in range(col_diff):
	series=[]
	column_index=alphabet[alphabet.index(col_start)+i]
	for j in range(row_diff):
		row_index=str(row_start+j)
		if j==0:
			series_names.append(ws[column_index+row_index].value)
		else:
			series.append(ws[column_index+row_index].value)
	data.append(series)
data=np.array(data)

#now, actually make the plot
plt.figure()
for i in range(col_diff):
	# if i==3:
	# 	plt.plot(data[i,:], color="C"+str(i),linestyle="--")
	# 	plt.plot(data[i,:1], color="C"+str(i),label=series_names[i])
	# else:
		plt.plot(data[i,:], color="C"+str(i),label=series_names[i])

	# if i==7:
	# 	plt.plot(data[i,:15], color="C"+str(i%4),linestyle="--")
	# 	plt.plot([14,15,16],data[i,14:], color="C"+str(i%4),linestyle="--",marker="v", markersize=2.5,markeredgecolor="black",markerfacecolor="black")
	# elif i==3:
	# 	plt.plot(data[i,:8], color="C"+str(i))
	# 	plt.plot([7,8,9,10,11,12,13,14],data[i,7:15], color="C"+str(i),linestyle="--")
	# 	plt.plot([14,15,16],data[i,14:], color="C"+str(i),linestyle="--",marker="^", markersize=2.5,markeredgecolor="black",markerfacecolor="black")
	# 	plt.plot(data[i,:1], color="C"+str(i),label=series_names[i])
	# elif i==6:
	# 	plt.plot(data[i,:14], color="C"+str(i%4),linestyle="--")
	# 	plt.plot([13,14],data[i,13:15], color="C"+str(i%4),)
	# 	plt.plot([14,15,16],data[i,14:], color="C"+str(i%4),marker="v", markersize=2.5,markeredgecolor="black",markerfacecolor="black")
	# elif i<4:
	# 	plt.plot(data[i,:15], color="C"+str(i),label=series_names[i])
	# 	plt.plot([14,15,16],data[i,14:], color="C"+str(i),marker="^", markersize=2.5,markeredgecolor="black",markerfacecolor="black")
	# else:
	# 	plt.plot(data[i,:15], color="C"+str(i%4))
	# 	plt.plot([14,15,16],data[i,14:], color="C"+str(i%4),marker="v", markersize=2.5,markeredgecolor="black",markerfacecolor="black")
plt.title(plot_title)
plt.legend(loc="lower left")
plt.ylabel('Orbital Energy (eV)')
plt.xlabel('Interpolation Coordinate')
plt.ylim(-0.6,1.2)
plt.grid(color='grey', linestyle='-', linewidth=.1)
plt.xticks(range(17),["Planar","","","","","","","","","","","","","","","","Pyramidal"])
# major_locator = MultipleLocator(base=1)
# plt.gca().yaxis.set_major_locator(major_locator)
# minor_locator = MultipleLocator(base=0.2)
# plt.gca().yaxis.set_minor_locator(minor_locator)
plt.savefig('plot_excel.pdf')
plt.show()
