import numpy as np
from pdos_helper import dos_grid_general,get_alpha,get_ao_ind, get_ind_ao_underc
from matplotlib import pyplot as plt
import random
rng = np.random.default_rng()
import pandas as pd
import sys
import math
import scipy.special
from openpyxl import Workbook, load_workbook
from scipy import stats

#A sandbox script meant to be used to plot things from the Defect Catalog, namely trap depths

file=sys.argv[1] #ideally the defect catalog itself, in my dropbox

#read in the defect catalog
wb = load_workbook(filename=file,data_only=True)
ws = wb.active

#which rows correspond to our target species?
target_rows=[]
for i in range(2000):
	if (ws["AK"+str(i+1)].value=="In") and (ws["AL"+str(i+1)].value==3):
		target_rows.append(str(i+1))

#now we extract desired parameters

pos_depths=[]
neg_depths=[]
neut_depths=[]
uc_depths=[]
no_uc_depths=[]
planar_depths=[]
pyramidal_depths=[]
a_count=0
b_count=0
c_count=0

in_angles=[]
ga_angles=[]
p_angles=[]
energies=[]

depth_cells=["BN","BO","BP","BQ","BR","BS","BT","BU","BV","BW","BX","BY","BZ","CA","CB","CC","CD","CE","CF","CG","CH","CI","CJ","CK","CL","CM","CN","CO","CP","CQ","CR","CS","CT","CU"]
for i, row in enumerate(target_rows):
		num_defects=int(ws["AL"+row].value)
		line=[]
		for j,col in enumerate(depth_cells):
			if (ws[col+row].value or ws[col+row].value==0):

				line.append(float(ws[col+row].value))

		if line !=[]:
			line=np.array(line)

			try:
				coplanarity=float(ws["BF"+row].value)
			except TypeError:
				print(row)
			try:
				dipole_overlap=float(ws["BJ"+row].value)
			except ValueError:
				print(row)
			uc_bound=ws["AP"+row].value
			num_f=int(ws["AN"+row].value)
			material=ws["G"+row].value
			cb_bulk=ws["V"+row].value
			vb_bulk=ws["R"+row].value
			if isinstance(ws["AD"+row].value,float):
				energy=ws["AD"+row].value
			charge=int(ws["H"+row].value)
			try:
				coordination=int(ws["AL"+row].value)
			except TypeError:
				print(row)
			# angles_3c=np.array([ws["AU"+row].value,ws["AV"+row].value,ws["AW"+row].value])
			# av_angle=np.mean(angles_3c)
			# p_angles.append(av_angle)

			# if coordination==2:
			# 	if isinstance(energy,float):
			# 		energies.append(energy)


			if material=="InP":
				a_count=a_count+num_defects
				in_angles.append(energy)
			# 	print(row)
				# angle_set=av_angle+0.266*rng.standard_normal(num_defects)
				# in_angles.extend(angle_set)
			elif material=="GaP":
				b_count=b_count+num_defects
				ga_angles.append(energy)
				# angle_set=av_angle+0.266*rng.standard_normal(num_defects)
				# ga_angles.extend(angle_set)
			# else:
			# 	c_count=c_count+num_defects

			
			for k,depth in enumerate(line):
						if k < num_defects:

							if coplanarity<.1:
								planar_depths.append(depth)
							elif coplanarity>=.1:
								pyramidal_depths.append(depth)


				
pos_depths=np.array(pos_depths)
neg_depths=np.array(neg_depths)
neut_depths=np.array(neut_depths)

# uc_depths=np.array(uc_depths)
# no_uc_depths=np.array(no_uc_depths)

# planar_depths=np.array(planar_depths)
# pyramidal_depths=np.array(pyramidal_depths)


in_angles=np.array(in_angles)
ga_angles=np.array(ga_angles)
# energies=np.array(energies)

print(np.average(in_angles), "InP P-2c Relaxed Relative Energy")
print(np.average(ga_angles), "GaP P-2c Relaxed Relative Energy")





E_grid = np.arange(0,3,0.001)

#plot the density of trap depths
sigma=0.1 # broadening parameter
# uc_grid=np.zeros(E_grid.shape)
# no_uc_grid=np.zeros(E_grid.shape)
# planar_grid=np.zeros(E_grid.shape)
# pyramidal_grid=np.zeros(E_grid.shape)
pos_grid=np.zeros(E_grid.shape)
neg_grid=np.zeros(E_grid.shape)
neut_grid=np.zeros(E_grid.shape)


for i in range(0,len(pos_depths)):
	pos_grid += np.exp(-(E_grid-pos_depths[i])**2/(2*sigma**2))
for i in range(0,len(neg_depths)):
	neg_grid += np.exp(-(E_grid-neg_depths[i])**2/(2*sigma**2))
for i in range(0,len(neut_depths)):
	neut_grid += np.exp(-(E_grid-neut_depths[i])**2/(2*sigma**2))
# for i in range(0,len(uc_depths)):
#     uc_grid += np.exp(-(E_grid-uc_depths[i])**2/(2*sigma**2))
# for i in range(0,len(no_uc_depths)):
# 	no_uc_grid += np.exp(-(E_grid-no_uc_depths[i])**2/(2*sigma**2))
# for i in range(0,len(planar_depths)):
#     planar_grid += np.exp(-(E_grid-planar_depths[i])**2/(2*sigma**2))
# for i in range(0,len(pyramidal_depths)):
#     pyramidal_grid += np.exp(-(E_grid-pyramidal_depths[i])**2/(2*sigma**2))



plt.figure()

# # plt.plot(E_grid,planar_grid+pyramidal_grid,color='C0',label='Total Ga-3c Trap Density')
# # plt.plot(E_grid,pyramidal_grid,color='C1',label='Pyramidal Ga-3c Trap Density')
# # plt.plot(E_grid,planar_grid,color='C2',label='Planar Ga-3c Trap Density')
# # plt.title("Pyramidal vs. Planar Ga-3c Trap Depths")

# plt.plot(E_grid,uc_grid+no_uc_grid,color='C0',label='Total P-3c Trap Density')
# plt.plot(E_grid,uc_grid,color='C1',label='P-3c bound to In/Ga-3c Trap Density')
# plt.plot(E_grid,no_uc_grid,color='C2',label='P-3c not bound to In/Ga-3c Trap Density')
# plt.title("Change in P-3c Trap Depth When Bound to In/Ga-3c")

# # plt.plot(E_grid,(uc_grid)+(no_uc_grid)+(pyramidal_grid)+(planar_grid),color='C0',label='Total Electron Trap Density')
# # plt.plot(E_grid,uc_grid,color='C1',label='In-3c Traps')
# # plt.plot(E_grid,no_uc_grid,color='C2',label='Structural In Traps')
# # plt.plot(E_grid,planar_grid,color='C3',label='Ga-3c Traps')
# # plt.plot(E_grid,pyramidal_grid,color='C4',label='Structural Ga Traps')
# # plt.title("Electron Trap Depth in InP vs. GaP")

# # plt.plot(E_grid,neg_grid+neut_grid+pos_grid,color='C0',label='Total P-3c Trap Density')
# # plt.plot(E_grid, neg_grid, color='C1', label="P-3c w/ Negative Dipole Overlap")
# # plt.plot(E_grid, neut_grid, color='C2', label="P-3c w/ Near Zero Dipole Overlap")
# # plt.plot(E_grid, pos_grid, color='C3', label="P-3c w/ Positive Dipole Overlap")
# # plt.title("Effect of Internal Dipole Moment on P-3c Trap Depths")

plt.plot(E_grid,neg_grid+neut_grid+pos_grid,color='C0',label='Total P-3c Trap Density')
plt.plot(E_grid, neg_grid, color='C1', label="Highly Bent P-3c")
plt.plot(E_grid, neut_grid, color='C2', label="Average Bend P-3c")
plt.plot(E_grid, pos_grid, color='C3', label="Highly Planar P-3c")
plt.title("Effect of Geometry on P-3c Trap Depths")

plt.legend()
plt.xlim(0,2)
plt.ylim(0,250)
plt.ylabel('Density of Trap Depths')
plt.xlabel('Trap Depth (eV)')
plt.legend(loc="upper right")

plt.savefig('p_3c_bend.pdf')
plt.show()



#