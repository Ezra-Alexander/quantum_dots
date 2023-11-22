import numpy as np
from pdos_helper import dos_grid_general,get_alpha,get_ao_ind, get_ind_ao_underc
from matplotlib import pyplot as plt
import random
import pandas as pd
import sys
import math
import scipy.special
from openpyxl import Workbook, load_workbook
from scipy import stats

#the goal of this script is to be able to make nice plots of the MO energies from interpolations
#reading in the excel files made by interpolation_analyser.sh

#this code assumes there are 17 frames in your interpolation

file=sys.argv[1] #the excel file to read data from
n_orbs=int(sys.argv[2]) #how many orbital energies do you want to plot, starting from the band edge. the excel file should only include occupied or virtual orbitals, so no need to specify
plot_title=sys.argv[3] #the title of the plot, in ""

#read in the data
wb = load_workbook(filename=file,data_only=True)
ws = wb.active

#extract target data into a n_orbs x n_frames array
alphabet="ABCDEFGHIJKLMNOPQRSTUVWXYZ" #lol
orb_energies=[]
for i in range(17):
	row=[]
	for j in range(n_orbs):
		row.append(ws[alphabet[j+2]+str(i+2)].value)
	orb_energies.append(row)

orb_energies=np.array(orb_energies)

#make the plot
plt.figure()
for i in range(n_orbs):
	if i==0:
		#plt.plot(orb_energies[:,i],color='C'+str(i),label="LUMO")
		plt.plot(orb_energies[:,i],color='C'+str(i),label="LUMO")
	elif i==1:
		#plt.plot(orb_energies[:,i],color='C'+str(i),label="LUMO+"+str(i))
		plt.plot(orb_energies[:,i],color='black', label="Other Ga MOs")
	else:
		plt.plot(orb_energies[:,i],color='black')
plt.legend(loc="upper right")
plt.title(plot_title)
plt.grid(color='grey', linestyle='-', linewidth=.1)
plt.ylim(min(orb_energies[:,0]-.6),max(orb_energies[:,n_orbs-1])+.4)
#plt.ylim(min(orb_energies[:,n_orbs-1])-.1,max(orb_energies[:,0]+.5))
#plt.ylim(min(orb_energies[:,n_orbs-1])-.1,-7.6)
plt.ylabel('Orbital Energy (eV)')
plt.xlabel('Interpolation Coordinate')
plt.xticks(range(17),["Bulk Crystal Geometry","","","","","","","","","","","","","","","","Distorted QD Geometry"])
plt.savefig('plot_interpolation.pdf')
plt.show()

