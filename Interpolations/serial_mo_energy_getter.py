import numpy as np
import sys
import math
from geom_helper import *
from openpyxl import Workbook, load_workbook

# for a set of serial single points, extracts specified MO energies and writes them to an excel file

out_file=sys.argv[1] #name of the serialized qchem.out
excel=sys.argv[2] #name of excel file to write
nval=int(sys.argv[3]) #number of orbitals to plot from band gap
occ=sys.argv[4] #either o for occupied or u for unoccupied

wb=Workbook()
ws = wb.active

#make column headers
alphabet="ABCDEFGHIJKLMNOPQRSTUVWXYZ" #lol
ws['A1']="Frame"
ws["B1"]="Overall Energy (eV)"
for i in range(nval):
	if occ=="o" or occ=="O":
		ws[alphabet[i+2]+"1"]="HOMO-"+str(i)+" Energy (eV)"
	elif occ=="u" or occ=="U":
		ws[alphabet[i+2]+"1"]="LUMO+"+str(i)+" Energy (eV)"
	else:
		raise Exception("Use 'o' for occupied orbitals or 'u' for unoccupied")

with open(out_file,"r") as out:
	e_tots=[]
	mo_e=[]
	mo_block=False
	for i,line in enumerate(out):
		if line.find('Welcome to Q-Chem')!= -1:
			frame_mos=[]
		if line.find('Total energy in the final basis set')!= -1:
			tot_e=line.strip().split()[-1]
			e_tots.append(float(tot_e)*27.211396641308)
		if line.find('Thank you very much for using Q-Chem.  Have a nice day.') != -1:
			if occ=="o" or occ=="O":
				e_val_edge=frame_mos[-nval:]
			elif occ=="u" or occ=="U":
				e_val_edge=frame_mos[:nval]
			ev_edge=[float(x)*27.211396641308 for x in e_val_edge]
			mo_e.append(ev_edge)
		if occ=="o" or occ=="O":
			if line.find('-- Occupied --')!= -1 and mo_block==False:
				mo_block=True
			elif mo_block==True and line.find('-- Virtual --')!= -1:
				mo_block=False
			elif mo_block==True:
				frame_mos.extend(line.strip().split())
		elif occ=="u" or occ=="U":
			if line.find('-- Virtual --')!= -1 and mo_block==False:
				mo_block=True
			elif mo_block==True and line.find('--------------------------------------------------------------')!= -1:
				mo_block=False
			elif mo_block==True:
				frame_mos.extend(line.strip().split())

for i,frame in enumerate(e_tots):
	ws["A"+str(i+2)]=i
	ws["B"+str(i+2)]=frame
	for j in range(nval):
		if occ=="o" or occ=="O":
			ws[alphabet[j+2]+str(i+2)]=mo_e[i][-(j+1)]
		elif occ=="u" or occ=="U":
			ws[alphabet[j+2]+str(i+2)]=mo_e[i][j]


wb.save(excel)