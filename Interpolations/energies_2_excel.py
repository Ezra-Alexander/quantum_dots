import numpy as np
import sys
import math
from geom_helper import *
from openpyxl import Workbook, load_workbook

#meant to be called by "interpolation_analyzer" in each subdirectory of a series of interpolation single points to add requested data to an excel file

#inputs
excel=sys.argv[1] #name of excel file to write
nval=int(sys.argv[2]) #number of orbitals to plot
first=int(sys.argv[3]) #what iteration you are on
frame=sys.argv[4] #which frame you are on
occ=sys.argv[5]

#if the notebook doesn't exist yet, make it. Otherwise, open it
if first==0:
	wb=Workbook()
else:
	wb = load_workbook(excel)

print("Processing", frame)

ws = wb.active

#unfortunately, bash considers "10" the second subdirectory in the ordering, so the "first" variable does not match which frame we are in. Thus, we extract the frame from the subdirectory we are in
frame_num=frame[:2]
if frame_num[1]=="t":
	frame_num=frame_num[0]
frame_num=int(frame_num)

#make column headers
alphabet="ABCDEFGHIJKLMNOPQRSTUVWXYZ" #lol
if first==0:
	ws['A1']="Frame"
	ws["B1"]="Overall Energy (eV)"
	for i in range(nval):
		if occ=="o" or occ=="O":
			ws[alphabet[i+2]+"1"]="HOMO-"+str(i)+" Energy (eV)"
		elif occ=="u" or occ=="U":
			ws[alphabet[i+2]+"1"]="LUMO+"+str(i)+" Energy (eV)"
		else:
			print("Use 'o' for occupied orbitals or 'u' for unoccupied")

#get the data
out_file="plot_"+str(frame_num)+".out"
e_tot=0
e_vals=[]
with open(out_file,"r") as out:
	e_test=False
	for i,line in enumerate(out):
		if occ=="o" or occ=="O":
			if line.find("Total energy")!=-1:
				e_tot=float(line.strip().split()[8])
			elif line.find("Alpha MOs")!=-1 and e_test==False:
				e_test=True
			elif e_test==True and line.find(" -- Occupied --")!=-1:
				print("Getting",nval,"highest occupied MO energies")
			elif e_test==True and (line.find(" -- Virtual --")!=-1 or line.find("Beta")!=-1):
				break
			elif e_test==True:
				e_vals.append(line.strip().split())
		elif occ=="u" or occ=="U":
			if line.find("Total energy")!=-1:
				e_tot=float(line.strip().split()[8])
			elif line.find("-- Virtual --")!=-1 and e_test==False:
				print("Getting",nval,"lowest unoccupied MO energies")
				e_test=True
			elif e_test==True and (line.find("--------------------------------------------------------------")!=-1 or line.find("Beta")!=-1):
				break
			elif e_test==True:
				e_vals.append(line.strip().split())


e_vals=sum(e_vals,[])

e_vals=[float(x) for x in e_vals]

if occ=="o" or occ=="O":
	e_val_edge=e_vals[-nval:]
elif occ=="u" or occ=="U":
	e_val_edge=e_vals[:nval]

e_tot=e_tot*27.211396641308 #convert to eV

ev_edge=[x*27.211396641308 for x in e_val_edge]

#write data to excel
ws["A"+str(frame_num+2)]=frame_num
ws["B"+str(frame_num+2)]=e_tot
for i in range(nval):
	if occ=="o" or occ=="O":
		ws[alphabet[i+2]+str(frame_num+2)]=ev_edge[-(i+1)]
	elif occ=="u" or occ=="U":
		ws[alphabet[i+2]+str(frame_num+2)]=ev_edge[i]


wb.save(excel)